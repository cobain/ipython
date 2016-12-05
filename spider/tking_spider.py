#coding:utf8
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import StaleElementReferenceException,TimeoutException
from time import sleep
import re
from openpyxl import load_workbook,Workbook
from xlrd import open_workbook
from xlutils.copy import copy
import subprocess
import sys
import os
from tqdm import *
import ConfigParser

reload(sys)
sys.setdefaultencoding('utf8')

class ticket_spider_niumowang(object):
    def __init__(self,url):
        self.browser = webdriver.PhantomJS('./phantomjs.exe')
        self.browser.get(url)

    def get_title_body(self):
        WebDriverWait(self.browser, 10).until(lambda the_driver: the_driver.find_element_by_class_name('type_img2').is_displayed())
        beautiful = BeautifulSoup(self.browser.page_source,'html.parser')
        ul_list = beautiful.findAll('ul')
        title = ul_list[2].contents[1].string
        begin_time = re.search(r'\d{4}\.\d{2}\.\d{2}',str(ul_list[2].contents[5])).group()
        try:
            end_time = re.search(r'\d{4}\.\d{2}\.\d{2}',str(ul_list[2].contents[5].contents[2])).group()
        except IndexError:
            end_time = ''
        place = re.search(r'class="ng-binding" id="jsvenue">(.*)<!-- ngIf:',str(ul_list[2].contents[7])).group(1)
        address = re.search(r'<!-- end ngIf: !!content.show.venueAddress -->(.*)</p>',str(ul_list[2].contents[7])).group(1)
        body_html = str(beautiful.select('#js_conxq')[0])
        # print body_html
        # body_html = re.sub(r'(^.*<img alt=".*?>).*$','\\1',body_html)
        imgs_list = []
        imgs_list.append(beautiful.select_one('.type_img2')['datasrc'])
        return title,begin_time,end_time,place,address,body_html,imgs_list

    def get_ticket_price(self):
        if self.browser.find_element_by_id('js_openbuy').text == u'缺货登记' or self.browser.find_element_by_id('js_openbuy').text == u'预约登记':
            return None
        self.browser.find_element_by_id('js_openbuy').click()
        try:
            WebDriverWait(self.browser, 10).until(lambda the_driver: the_driver.find_element_by_xpath('//ul[@ng-if="!content.isPermanent"]').is_displayed())
            day_elements_list = self.browser.find_elements_by_xpath('//ul[@ng-if="!content.isPermanent"]')
        except TimeoutException:
            WebDriverWait(self.browser, 10).until(lambda the_driver: the_driver.find_element_by_xpath('//ul[@ng-if="!!content.isPermanent"]').is_displayed())
            day_elements_list = self.browser.find_elements_by_xpath('//ul[@ng-if="!!content.isPermanent"]')
        try:
            WebDriverWait(self.browser, 10).until(lambda the_driver: the_driver.find_element_by_xpath('//ul[@ng-if="!!sp.comments"]').is_displayed())
            price_elements_list = self.browser.find_elements_by_xpath('//ul[@ng-if="!!sp.comments"]')
        except TimeoutException:
            WebDriverWait(self.browser, 10).until(lambda the_driver: the_driver.find_element_by_xpath('//ul[@ng-if="!sp.comments"]').is_displayed())
            price_elements_list = self.browser.find_elements_by_xpath('//ul[@ng-if="!sp.comments"]')
        ticket_price_list = []
        for i in range(len(day_elements_list)):
            if day_elements_list[i].get_attribute('disabled'):
                continue
            day_elements_list[i].click()
            sleep(0.1)
            price_elements_list = self.browser.find_elements_by_xpath('//div[@ng-disabled="!sp.available"]')
            for j in range(len(price_elements_list)):
                if price_elements_list[j].get_attribute('disabled'):
                    continue
                price_elements_list[j].click()
                sleep(0.1)
                ticket_price_list.append(day_elements_list[i].find_element_by_tag_name('a').text + '\n' + price_elements_list[j].find_element_by_tag_name('a').text \
                        + '\n' + self.browser.find_element_by_id('js_book_sj').text + u'元')
        return ticket_price_list

    def close_browser(self):
        self.browser.close()

def get_url_list():
    url_list = []
    browser = webdriver.PhantomJS('./phantomjs.exe')
    browser.get('http://www.tking.cn/list/1#nmwz')
    WebDriverWait(browser, 10).until(lambda the_driver: the_driver.find_element_by_link_text('末页').is_displayed())
    browser.find_element_by_link_text('末页').click()
    last_page_id = browser.find_elements_by_tag_name('ul')[-3].find_elements_by_tag_name('a')[-3].text
    browser.close()
    print 'start get urls...'
    sleep(0.1)
    for i in tqdm(range(1,int(last_page_id)+1)):
        for j in range(10):
            while 1:
                try:
                    browser = webdriver.PhantomJS('./phantomjs.exe')
                    browser.get('http://www.tking.cn/list/1#nmwz')
                    for page_id in range(i-1):
                        WebDriverWait(browser, 10).until(lambda the_driver: the_driver.find_element_by_class_name('type_img3').is_displayed())
                        reference = browser.find_elements_by_class_name('type_img3')[0]
                        browser.find_element_by_link_text('下一页').click()
                        WebDriverWait(browser, 10).until(lambda the_driver: the_driver.find_elements_by_class_name('type_img3')[0] != reference)
                    try:
                        WebDriverWait(browser, 5).until(lambda the_driver: the_driver.find_elements_by_class_name('type_img3')[j].is_displayed())
                        browser.find_elements_by_class_name('type_img3')[j].click()
                    except IndexError:
                        break
                    browser.switch_to_window(browser.window_handles[1])
                    WebDriverWait(browser, 10).until(lambda the_driver: the_driver.find_element_by_class_name('type_img2').is_displayed())
                    url_list.append(browser.current_url)
                    browser.close()
                    browser.switch_to_window(browser.window_handles[0])
                    browser.close()
                    break
                except StaleElementReferenceException:
                    continue
    return url_list

def create_excel(data_list):
    wb = Workbook()
    ws0 = wb.active
    wb.remove_sheet(ws0)
    ws1 = wb.create_sheet('Sheet1')
    ws2 = wb.create_sheet('Sheet2')

    ws1['A1'] = 'id'
    ws1['B1'] = 'eventname'
    ws1['C1'] = 'eventlocation'
    ws1['D1'] = 'eventvenue'
    ws1['E1'] = 'eventpictures'
    ws1['F1'] = 'eventwebdescription'
    ws1['G1'] = 'eventstartdatetime'
    ws1['H1'] = 'eventenddatetime'
    ws1['I1'] = 'index'
    ws1['J1'] = 'express_charge'
    ws2['A1'] = 'id'
    ws2['B1'] = 'subeventtitle'
    ws2['C1'] = 'subeventdescription'
    ws2['D1'] = 'subeventprice'
    ws2['E1'] = 'subeventoriginalprice'
    ws2['F1'] = 'goodsnum'
    ws2['G1'] = 'singleorder_maxcount'
    ws2['H1'] = 'singleuser_maxcount'
    ws2['I1'] = 'subevent_attributes'
    ws1_current_row = 2
    ws2_current_row = 2
    for _id in range(1,len(data_list)+1):
        ws1['A'+str(ws1_current_row)] = _id
        ws1['B'+str(ws1_current_row)] = data_list[_id-1]['title']
        ws1['C'+str(ws1_current_row)] = data_list[_id-1]['address']
        ws1['D'+str(ws1_current_row)] = data_list[_id-1]['place']
        ws1['E'+str(ws1_current_row)] = ','.join(data_list[_id-1]['imgs_list'])
        ws1['F'+str(ws1_current_row)] = data_list[_id-1]['body_html']
        ws1['G'+str(ws1_current_row)] = data_list[_id-1]['begin_time']
        if data_list[_id-1]['end_time'] == '':
            ws1['H'+str(ws1_current_row)] = data_list[_id-1]['begin_time']
        else:
            ws1['H'+str(ws1_current_row)] = data_list[_id-1]['end_time']
        ws1['I'+str(ws1_current_row)] = data_list[_id-1]['index']
        ws1['J'+str(ws1_current_row)] = 12
        if data_list[_id-1]['ticket_price_list'] != None:
            for i in data_list[_id-1]['ticket_price_list']:
                ticket_info_list = i.split('\n')
                if len(ticket_info_list) == 4:
                    ws2['A'+str(ws2_current_row)] = _id
                    ws2['B'+str(ws2_current_row)] = '{0} {1} {2}'.format(ticket_info_list[0],ticket_info_list[1],ticket_info_list[2])
                    ws2['C'+str(ws2_current_row)] = ''
                    ws2['D'+str(ws2_current_row)] = re.search(r'\d+',ticket_info_list[3]).group()
                    ws2['E'+str(ws2_current_row)] = re.search(r'\d+',ticket_info_list[2]).group()
                    ws2_current_row += 1
                else:
                    ws2['A'+str(ws2_current_row)] = _id
                    ws2['B'+str(ws2_current_row)] = '{0} {1} {2} {3}'.format(ticket_info_list[0],ticket_info_list[1],ticket_info_list[2],ticket_info_list[3])
                    ws2['C'+str(ws2_current_row)] = ticket_info_list[2]
                    ws2['D'+str(ws2_current_row)] = re.search(r'\d+',ticket_info_list[4]).group()
                    ws2['E'+str(ws2_current_row)] = re.search(r'\d+',ticket_info_list[3]).group()
                    ws2_current_row += 1
            ws1_current_row += 1
    wb.save('result.xlsx')

    rb = open_workbook('result.xlsx')
    wb = copy(rb)
    wb.save('result.xls')
    os.remove('result.xlsx')

if __name__ == '__main__':
    config = ConfigParser.ConfigParser()
    config.read('settings.ini')
    fetch_all = config.get('config','fetch_all')
    url_dir = config.get('config','url_dir')
    if fetch_all == 'yes':
        url_list = get_url_list()
        print len(url_list)
    else:
        with open(url_dir) as f:
            url_list = f.readlines()
        while '' in url_list:
            url_list.remove('')
    data_list = []
    # url_list = ['http://www.tking.cn/content/5719e2720cf22ad1941adbb3']
    # url_list = ['http://www.tking.cn/content/5702535a0cf289d08a11203d',
    #             'http://www.tking.cn/content/572c70170cf2ce8d3a6e4a08',
    #             'http://www.tking.cn/content/573ebde30cf2626ce904d490',
    #             'http://www.tking.cn/content/572029e80cf271ce13ceb00e',
    #             'http://www.tking.cn/content/57610c030cf28969e9163c58',
    #             'http://www.tking.cn/content/57202d570cf271ce13ceb01c',
    #             'http://www.tking.cn/content/5720203e0cf271ce13ceaff2',
    #             'http://www.tking.cn/content/56ced5e30cf252bc2e05dd00',
    #             'http://www.tking.cn/content/5785b4330cf2c312d7db69c0',
    #             'http://www.tking.cn/content/5719e2720cf22ad1941adbb3'
    #             ]
    # url_list = get_url_list()
    print
    print 'start get page info...'
    sleep(0.1)
    for _id in tqdm(range(1,len(url_list)+1)):
        # print 'url {0} start'.format(_id)
        info = ticket_spider_niumowang(url_list[_id-1])
        title,begin_time,end_time,place,address,body_html,imgs_list = info.get_title_body()
        ticket_price_list = info.get_ticket_price()
        info.close_browser()
        data_list.append({'title':title,
                          'begin_time':begin_time,
                          'end_time':end_time,
                          'place':place,
                          'address':address,
                          'body_html':body_html,
                          'imgs_list':imgs_list,
                          'ticket_price_list':ticket_price_list,
                          'index':os.path.basename(url_list[_id-1])})
    create_excel(data_list)
    print 'all done'
    # subprocess.Popen('taskkill -im chromedriver.exe -f ',shell=True,stdout=subprocess.PIPE)
